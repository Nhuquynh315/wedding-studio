import csv
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timezone
from datetime import date as date_type

from flask import (Blueprint, Response, flash, jsonify, redirect,
                   render_template, request, session as flask_session, url_for)
from flask_login import login_required

from app import db
from app.models import BudgetCategory, Expense, Vendor, VENDOR_CATEGORIES, VENDOR_STATUSES
from app.routes.utils import get_vendor_or_403, get_wedding_or_403
from app.routes.budget import find_matching_category

vendors_bp = Blueprint('vendors', __name__)

# Map vendor category → closest budget category name (for auto-expense)
_CAT_MAP = {
    'Venue':        'Venue & Catering',
    'Catering':     'Venue & Catering',
    'Photography':  'Photography & Video',
    'Videography':  'Photography & Video',
    'Flowers':      'Flowers & Decor',
    'Music':        'Music & Entertainment',
    'Hair & Makeup':'Beauty',
    'Transport':    'Transport',
    'Cake':         'Catering & Cake',
    'Stationery':   'Stationery',
    'Officiant':    'Other',
    'Other':        'Other',
}


def _utcnow():
    return datetime.now(timezone.utc)


def _parse_date(raw):
    if raw:
        try:
            return date_type.fromisoformat(raw.strip())
        except ValueError:
            pass
    return None


def _parse_float(raw):
    if raw:
        try:
            return float(raw.strip())
        except ValueError:
            pass
    return None


def _vendor_summary(wedding):
    vendors = wedding.vendors
    total        = len(vendors)
    booked       = sum(1 for v in vendors if v.status == 'booked')
    pending      = sum(1 for v in vendors if v.status in ('considering', 'backup'))
    total_quoted = sum(v.quoted_price or 0 for v in vendors if v.status == 'booked')
    deposits_paid = sum(1 for v in vendors if v.deposit_paid)
    return dict(total=total, booked=booked, pending=pending,
                total_quoted=total_quoted, deposits_paid=deposits_paid)


# ── Main vendor page ───────────────────────────────────────────────────
@vendors_bp.route('/wedding/<int:wedding_id>/vendors')
@login_required
def vendors(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    all_vendors = (Vendor.query
                   .filter_by(wedding_id=wedding_id)
                   .order_by(Vendor.category, Vendor.business_name)
                   .all())
    # Tabs: use actual distinct category values from vendors (preserves order)
    seen = set()
    used_categories = []
    for v in all_vendors:
        if v.category not in seen:
            seen.add(v.category)
            used_categories.append(v.category)
    used_categories.sort()

    summary = _vendor_summary(wedding)
    budget_categories = (BudgetCategory.query
                         .filter_by(wedding_id=wedding_id)
                         .order_by(BudgetCategory.name)
                         .all())

    # Map vendor_id → linked expense (for budget link display)
    linked_expenses = {
        e.vendor_id: e
        for e in Expense.query.filter_by(wedding_id=wedding_id)
                               .filter(Expense.vendor_id.isnot(None)).all()
    }

    # Backfill: silently create missing budget expenses for already-booked vendors
    if budget_categories:
        needs_commit = False
        for v in all_vendors:
            if v.status == 'booked' and v.quoted_price and v.id not in linked_expenses:
                cat = find_matching_category(wedding_id, v.category)
                expense = Expense(
                    wedding_id     = wedding_id,
                    category_id    = cat.id if cat else None,
                    vendor_id      = v.id,
                    title          = v.business_name,
                    estimated_cost = v.quoted_price,
                    is_paid        = v.deposit_paid,
                    actual_cost    = v.deposit_amount if v.deposit_paid else None,
                    paid_date      = date_type.today() if v.deposit_paid else None,
                )
                db.session.add(expense)
                linked_expenses[v.id] = expense
                needs_commit = True
        if needs_commit:
            try:
                db.session.commit()
                # Re-fetch so expenses have IDs and category relationships loaded
                linked_expenses = {
                    e.vendor_id: e
                    for e in Expense.query.filter_by(wedding_id=wedding_id)
                                           .filter(Expense.vendor_id.isnot(None)).all()
                }
            except Exception:
                db.session.rollback()

    # Category options for the add/edit dropdowns
    # If budget categories exist, use those names so vendor ↔ budget categories match exactly.
    # Otherwise fall back to the hardcoded list.
    form_categories = (
        [bc.name for bc in budget_categories] if budget_categories else list(VENDOR_CATEGORIES)
    )

    # Connection 5: pick up any pending task suggestion from add_vendor
    task_suggestion = flask_session.pop('task_suggestion', None)

    return render_template(
        'wedding/vendors.html',
        wedding=wedding,
        all_vendors=all_vendors,
        used_categories=used_categories,
        vendor_categories=form_categories,
        vendor_statuses=VENDOR_STATUSES,
        summary=summary,
        budget_categories=budget_categories,
        linked_expenses=linked_expenses,
        has_budget=len(budget_categories) > 0,
        date_today=date_type.today(),
        task_suggestion=task_suggestion,
    )


# ── Add vendor ─────────────────────────────────────────────────────────
@vendors_bp.route('/wedding/<int:wedding_id>/vendors/add', methods=['POST'])
@login_required
def add_vendor(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    business_name = request.form.get('business_name', '').strip()
    if not business_name:
        flash('Business name is required.', 'danger')
        return redirect(url_for('vendors.vendors', wedding_id=wedding_id))

    category = request.form.get('category', 'Other').strip() or 'Other'

    quoted   = _parse_float(request.form.get('quoted_price'))
    deposit  = _parse_float(request.form.get('deposit_amount'))
    balance  = (quoted - deposit) if (quoted and deposit and quoted > deposit) else None

    vendor = Vendor(
        wedding_id             = wedding_id,
        category               = category,
        business_name          = business_name,
        contact_name           = request.form.get('contact_name', '').strip() or None,
        email                  = request.form.get('email', '').strip() or None,
        phone                  = request.form.get('phone', '').strip() or None,
        website                = request.form.get('website', '').strip() or None,
        quoted_price           = quoted,
        deposit_amount         = deposit,
        deposit_due_date       = _parse_date(request.form.get('deposit_due_date')),
        final_payment_amount   = balance,
        final_payment_due_date = _parse_date(request.form.get('final_payment_due_date')),
        rating                 = int(request.form.get('rating') or 0) or None,
        notes                  = request.form.get('notes', '').strip() or None,
        status                 = request.form.get('status', 'considering'),
    )
    db.session.add(vendor)
    try:
        db.session.commit()
        flash(f'"{business_name}" added.', 'success')
        # Connection 5: suggest a checklist task if none exists for this vendor category
        from app.utils.connections import VENDOR_TO_TASK
        from app.models import ChecklistItem
        keywords = VENDOR_TO_TASK.get(vendor.category, [])
        if keywords:
            existing_items = ChecklistItem.query.filter_by(wedding_id=wedding_id).all()
            has_task = any(
                any(kw in item.title.lower() for kw in keywords)
                for item in existing_items
            )
            if not has_task:
                flask_session['task_suggestion'] = {
                    'title': f'Book {vendor.category.lower()}',
                    'wedding_id': wedding_id,
                    'category': vendor.category,
                }
    except Exception:
        db.session.rollback()
        flash('Could not add vendor.', 'danger')

    anchor = category.lower().replace(' ', '-').replace('&', '').replace('--', '-')
    return redirect(url_for('vendors.vendors', wedding_id=wedding_id) + f'#{anchor}')


# ── Edit vendor ────────────────────────────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/edit', methods=['POST'])
@login_required
def edit_vendor(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    business_name = request.form.get('business_name', '').strip()
    if not business_name:
        flash('Business name is required.', 'danger')
        return redirect(url_for('vendors.vendors', wedding_id=vendor.wedding_id))

    category = request.form.get('category', vendor.category).strip() or vendor.category

    quoted  = _parse_float(request.form.get('quoted_price'))
    deposit = _parse_float(request.form.get('deposit_amount'))
    balance = (quoted - deposit) if (quoted and deposit and quoted > deposit) else None

    vendor.business_name           = business_name
    vendor.category                = category
    vendor.contact_name            = request.form.get('contact_name', '').strip() or None
    vendor.email                   = request.form.get('email', '').strip() or None
    vendor.phone                   = request.form.get('phone', '').strip() or None
    vendor.website                 = request.form.get('website', '').strip() or None
    vendor.quoted_price            = quoted
    vendor.deposit_amount          = deposit
    vendor.deposit_due_date        = _parse_date(request.form.get('deposit_due_date'))
    vendor.final_payment_amount    = balance
    vendor.final_payment_due_date  = _parse_date(request.form.get('final_payment_due_date'))
    vendor.rating                  = int(request.form.get('rating') or 0) or None
    vendor.notes                   = request.form.get('notes', '').strip() or None
    vendor.status                  = request.form.get('status', vendor.status)

    try:
        db.session.commit()
        flash('Vendor updated.', 'success')
    except Exception:
        db.session.rollback()
        flash('Could not update vendor.', 'danger')
    return redirect(url_for('vendors.vendors', wedding_id=vendor.wedding_id))


# ── Delete vendor ──────────────────────────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/delete', methods=['POST'])
@login_required
def delete_vendor(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    wedding_id = vendor.wedding_id
    db.session.delete(vendor)
    try:
        db.session.commit()
        flash('Vendor deleted.', 'success')
    except Exception:
        db.session.rollback()
        flash('Could not delete vendor.', 'danger')
    return redirect(url_for('vendors.vendors', wedding_id=wedding_id))


# ── Update status (AJAX) ───────────────────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/status', methods=['POST'])
@login_required
def update_status(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    new_status = (request.json or {}).get('status') or request.form.get('status')
    if new_status not in VENDOR_STATUSES:
        return jsonify({'ok': False, 'error': 'Invalid status'}), 400

    old_status = vendor.status
    vendor.status = new_status

    budget_updated = False
    category_name = None

    # Auto-create budget expense when marking as booked (and none exists yet)
    if new_status == 'booked' and old_status != 'booked' and vendor.quoted_price:
        existing = Expense.query.filter_by(vendor_id=vendor.id).first()
        if not existing:
            cat = find_matching_category(vendor.wedding_id, vendor.category)
            expense = Expense(
                wedding_id     = vendor.wedding_id,
                category_id    = cat.id if cat else None,
                vendor_id      = vendor.id,
                title          = vendor.business_name,
                estimated_cost = vendor.quoted_price,
                is_paid        = False,
            )
            db.session.add(expense)
            budget_updated = True
            category_name = cat.name if cat else None

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'DB error'}), 500

    # Connection 1: auto-complete matching checklist task when vendor is booked
    task_completed = None
    if new_status == 'booked' and old_status != 'booked':
        from app.utils.connections import auto_complete_vendor_task
        task_completed = auto_complete_vendor_task(vendor.wedding_id, vendor.category)

    summary = _vendor_summary(vendor.wedding)
    return jsonify({
        'ok': True,
        'status': new_status,
        'budget_updated': budget_updated,
        'category_name': category_name,
        'vendor_id': vendor.id,
        'summary': summary,
        'task_completed': task_completed,
    })


# ── Toggle deposit paid (AJAX) ─────────────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/toggle-deposit', methods=['POST'])
@login_required
def toggle_deposit(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    vendor.deposit_paid = not vendor.deposit_paid

    # Sync the linked budget expense (if any)
    linked = Expense.query.filter_by(vendor_id=vendor.id).first()
    if linked:
        if vendor.deposit_paid:
            linked.is_paid     = True
            linked.actual_cost = vendor.deposit_amount
            linked.paid_date   = date_type.today()
        else:
            linked.is_paid     = False
            linked.actual_cost = None
            linked.paid_date   = None

    try:
        db.session.commit()
        summary = _vendor_summary(vendor.wedding)
        return jsonify({'ok': True, 'deposit_paid': vendor.deposit_paid, 'summary': summary})
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False}), 500


# ── Toggle final payment (AJAX) ───────────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/toggle-final-payment', methods=['POST'])
@login_required
def toggle_final_payment(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    vendor.final_payment_paid = not vendor.final_payment_paid

    # Sync the linked budget expense
    linked = Expense.query.filter_by(vendor_id=vendor.id).first()
    if linked:
        if vendor.final_payment_paid:
            linked.is_paid     = True
            linked.actual_cost = vendor.quoted_price
            linked.paid_date   = date_type.today()
        else:
            # Revert to deposit-only paid state
            linked.is_paid     = vendor.deposit_paid
            linked.actual_cost = vendor.deposit_amount if vendor.deposit_paid else None
            linked.paid_date   = date_type.today() if vendor.deposit_paid else None

    try:
        db.session.commit()
        summary = _vendor_summary(vendor.wedding)
        return jsonify({'ok': True, 'final_payment_paid': vendor.final_payment_paid, 'summary': summary})
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False}), 500


# ── Toggle contracted (AJAX) ───────────────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/toggle-contracted', methods=['POST'])
@login_required
def toggle_contracted(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    vendor.contracted = not vendor.contracted
    vendor.contract_signed_date = date_type.today() if vendor.contracted else None
    try:
        db.session.commit()
        return jsonify({'ok': True, 'contracted': vendor.contracted,
                        'signed_date': vendor.contract_signed_date.isoformat() if vendor.contract_signed_date else None})
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False}), 500


# ── Auto-create budget expense (AJAX) ─────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/add-to-budget', methods=['POST'])
@login_required
def add_to_budget(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    if not vendor.quoted_price:
        return jsonify({'ok': False, 'error': 'No quoted price'}), 400

    # Find the best matching budget category
    preferred_name = _CAT_MAP.get(vendor.category, 'Other')
    cats = BudgetCategory.query.filter_by(wedding_id=vendor.wedding_id).all()
    cat = next((c for c in cats if c.name == preferred_name), None)
    if cat is None and cats:
        # fuzzy: partial match on category keyword
        keyword = vendor.category.split()[0].lower()
        cat = next((c for c in cats if keyword in c.name.lower()), cats[0])

    expense = Expense(
        wedding_id     = vendor.wedding_id,
        category_id    = cat.id if cat else None,
        vendor_id      = vendor.id,
        title          = vendor.business_name,
        estimated_cost = vendor.quoted_price,
    )
    db.session.add(expense)
    try:
        db.session.commit()
        return jsonify({'ok': True, 'expense_id': expense.id,
                        'category': cat.name if cat else None})
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'DB error'}), 500


# ── Update rating (AJAX) ───────────────────────────────────────────────
@vendors_bp.route('/vendor/<int:vendor_id>/rating', methods=['POST'])
@login_required
def update_rating(vendor_id):
    vendor = get_vendor_or_403(vendor_id)
    try:
        rating = int((request.json or {}).get('rating', 0))
        if not 1 <= rating <= 5:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Rating must be 1–5'}), 400

    vendor.rating = rating
    try:
        db.session.commit()
        return jsonify({'ok': True, 'rating': rating})
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False}), 500


# ── CSV export ─────────────────────────────────────────────────────────
@vendors_bp.route('/wedding/<int:wedding_id>/vendors/export')
@login_required
def export_vendors(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    vendors_list = (Vendor.query
                    .filter_by(wedding_id=wedding_id)
                    .order_by(Vendor.category, Vendor.business_name)
                    .all())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Category', 'Business Name', 'Contact Name', 'Email', 'Phone',
        'Website', 'Status', 'Quoted Price', 'Deposit Amount', 'Deposit Paid',
        'Deposit Due Date', 'Contracted', 'Contract Signed Date', 'Rating', 'Notes',
    ])
    for v in vendors_list:
        writer.writerow([
            v.category, v.business_name, v.contact_name or '', v.email or '',
            v.phone or '', v.website or '', v.status,
            v.quoted_price or '', v.deposit_amount or '', 'Yes' if v.deposit_paid else 'No',
            v.deposit_due_date.isoformat() if v.deposit_due_date else '',
            'Yes' if v.contracted else 'No',
            v.contract_signed_date.isoformat() if v.contract_signed_date else '',
            v.rating or '', v.notes or '',
        ])

    filename = f"{wedding.partner1_name}_{wedding.partner2_name}_vendors.csv".replace(' ', '_')
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ── Excel export ────────────────────────────────────────────────────────
@vendors_bp.route('/wedding/<int:wedding_id>/vendors/export/excel')
@login_required
def export_vendors_excel(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    vendors_list = (Vendor.query
                    .filter_by(wedding_id=wedding_id)
                    .order_by(Vendor.category, Vendor.business_name)
                    .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vendors'

    headers = [
        'Category', 'Business Name', 'Contact Name', 'Email', 'Phone',
        'Website', 'Status', 'Quoted Price', 'Deposit Amount', 'Deposit Paid',
        'Deposit Due Date', 'Contracted', 'Contract Signed Date', 'Rating', 'Notes',
    ]

    header_fill = PatternFill(start_color='C9687A', end_color='C9687A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for v in vendors_list:
        ws.append([
            v.category, v.business_name, v.contact_name or '', v.email or '',
            v.phone or '', v.website or '', v.status,
            v.quoted_price or '', v.deposit_amount or '', 'Yes' if v.deposit_paid else 'No',
            v.deposit_due_date.isoformat() if v.deposit_due_date else '',
            'Yes' if v.contracted else 'No',
            v.contract_signed_date.isoformat() if v.contract_signed_date else '',
            v.rating or '', v.notes or '',
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{wedding.partner1_name}_{wedding.partner2_name}_vendors.xlsx".replace(' ', '_')
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
