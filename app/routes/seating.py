import csv
import io

import openpyxl
from openpyxl.styles import Font, PatternFill
from flask import Blueprint, jsonify, render_template, request, abort, make_response
from flask_login import login_required, current_user

from app import db
from app.models import Guest, WeddingTable
from app.routes.utils import get_wedding_or_403, get_guest_or_403, get_table_or_403

seating_bp = Blueprint('seating', __name__)


@seating_bp.route('/wedding/<int:wedding_id>/seating')
@login_required
def seating(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    tables = (WeddingTable.query
              .filter_by(wedding_id=wedding_id)
              .order_by(WeddingTable.table_number)
              .all())
    unassigned = (Guest.query
                  .filter_by(wedding_id=wedding_id, rsvp_status='confirmed')
                  .filter(Guest.table_id.is_(None))
                  .order_by(Guest.full_name)
                  .all())
    accepted_count = (Guest.query
                      .filter_by(wedding_id=wedding_id, rsvp_status='confirmed')
                      .count())
    accepted_guests = (Guest.query
                       .filter_by(wedding_id=wedding_id, rsvp_status='confirmed')
                       .order_by(Guest.full_name)
                       .all())
    raw_groups = (db.session.query(Guest.group_name)
                  .filter_by(wedding_id=wedding_id, rsvp_status='confirmed')
                  .filter(Guest.group_name.isnot(None))
                  .filter(Guest.group_name != '')
                  .distinct()
                  .all())
    groups = sorted([row[0] for row in raw_groups])
    return render_template(
        'wedding/seating.html',
        wedding=wedding,
        tables=tables,
        unassigned=unassigned,
        accepted_count=accepted_count,
        accepted_guests=accepted_guests,
        groups=groups,
    )


@seating_bp.route('/wedding/<int:wedding_id>/seating/tables/add', methods=['POST'])
@login_required
def add_table(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    shape = request.form.get('shape', 'round')
    if shape not in ('round', 'rectangle', 'head'):
        shape = 'round'
    existing_count = WeddingTable.query.filter_by(wedding_id=wedding_id).count()
    max_num_row = (db.session.query(db.func.max(WeddingTable.table_number))
                   .filter_by(wedding_id=wedding_id)
                   .scalar())
    next_number = (max_num_row or 0) + 1
    col = existing_count % 4
    row = existing_count // 4
    pos_x = 80 + col * 250
    pos_y = 80 + row * 260
    table = WeddingTable(
        wedding_id=wedding_id,
        table_number=next_number,
        capacity=8,
        shape=shape,
        position_x=pos_x,
        position_y=pos_y,
    )
    try:
        db.session.add(table)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    return jsonify({
        'ok': True,
        'id': table.id,
        'table_number': table.table_number,
        'table_name': table.table_name,
        'capacity': table.capacity,
        'shape': table.shape,
        'position_x': table.position_x,
        'position_y': table.position_y,
        'display_name': table.display_name(),
    })


@seating_bp.route('/seating/table/<int:table_id>/edit', methods=['POST'])
@login_required
def edit_table(table_id):
    table = get_table_or_403(table_id)
    table.table_name = request.form.get('table_name', '').strip() or None
    try:
        table.capacity = int(request.form.get('capacity', table.capacity))
    except (ValueError, TypeError):
        pass
    shape = request.form.get('shape', table.shape)
    if shape in ('round', 'rectangle', 'head'):
        table.shape = shape
    table.notes = request.form.get('notes', '').strip() or None
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    return jsonify({
        'ok': True,
        'table_name': table.table_name,
        'display_name': table.display_name(),
        'capacity': table.capacity,
        'shape': table.shape,
        'notes': table.notes or '',
    })


@seating_bp.route('/seating/table/<int:table_id>/delete', methods=['POST'])
@login_required
def delete_table(table_id):
    table = get_table_or_403(table_id)
    for guest in list(table.guests):
        guest.table_id = None
        guest.table_number = None
    try:
        db.session.delete(table)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    return jsonify({'ok': True})


@seating_bp.route('/seating/table/<int:table_id>/position', methods=['POST'])
@login_required
def update_position(table_id):
    table = get_table_or_403(table_id)
    data = request.get_json(force=True, silent=True) or {}
    try:
        table.position_x = float(data.get('x', table.position_x or 0))
        table.position_y = float(data.get('y', table.position_y or 0))
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Invalid coordinates'}), 400
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    return jsonify({'ok': True})


@seating_bp.route('/seating/assign', methods=['POST'])
@login_required
def assign_guest():
    data = request.get_json(force=True, silent=True) or {}
    guest_id = data.get('guest_id')
    table_id = data.get('table_id')
    if not guest_id or not table_id:
        return jsonify({'ok': False, 'error': 'Missing guest_id or table_id'}), 400
    guest = get_guest_or_403(guest_id)
    table = get_table_or_403(table_id)
    if guest.wedding_id != table.wedding_id:
        abort(403)
    if len(table.guests) >= table.capacity:
        return jsonify({'ok': False, 'error': 'Table is at capacity'}), 400
    guest.table_id = table.id
    guest.table_number = table.table_number
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    return jsonify({'ok': True, 'count': len(table.guests)})


@seating_bp.route('/seating/unassign', methods=['POST'])
@login_required
def unassign_guest():
    data = request.get_json(force=True, silent=True) or {}
    guest_id = data.get('guest_id')
    if not guest_id:
        return jsonify({'ok': False, 'error': 'Missing guest_id'}), 400
    guest = get_guest_or_403(guest_id)
    guest.table_id = None
    guest.table_number = None
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    return jsonify({'ok': True})


@seating_bp.route('/wedding/<int:wedding_id>/seating/auto-assign', methods=['POST'])
@login_required
def auto_assign(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    tables = (WeddingTable.query
              .filter_by(wedding_id=wedding_id)
              .order_by(WeddingTable.table_number)
              .all())
    unassigned = (Guest.query
                  .filter_by(wedding_id=wedding_id, rsvp_status='confirmed')
                  .filter(Guest.table_id.is_(None))
                  .all())
    # Sort guests: group together, then by name within group
    unassigned_sorted = sorted(
        unassigned,
        key=lambda g: (g.group_name or 'zzz', g.full_name)
    )
    # Track counts in a local dict — relationship cache doesn't update mid-loop
    counts = {t.id: len(t.guests) for t in tables}
    assigned_count = 0
    failed_count = 0
    for guest in unassigned_sorted:
        placed = False
        for table in tables:
            if counts[table.id] < table.capacity:
                guest.table_id = table.id
                guest.table_number = table.table_number
                counts[table.id] += 1
                placed = True
                assigned_count += 1
                break
        if not placed:
            failed_count += 1
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    if failed_count > 0:
        message = f'Assigned {assigned_count} guests. {failed_count} could not be placed (tables full).'
    else:
        message = f'Successfully assigned {assigned_count} guests to tables.'
    return jsonify({
        'ok': True,
        'message': message,
        'assigned': assigned_count,
        'failed': failed_count,
    })


@seating_bp.route('/wedding/<int:wedding_id>/seating/export')
@login_required
def export_csv(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    fmt = request.args.get('format', 'csv')
    tables = (WeddingTable.query
              .filter_by(wedding_id=wedding_id)
              .order_by(WeddingTable.table_number)
              .all())
    unassigned = (Guest.query
                  .filter_by(wedding_id=wedding_id, rsvp_status='confirmed')
                  .filter(Guest.table_id.is_(None))
                  .order_by(Guest.full_name)
                  .all())

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Seating Chart'
        headers = ['Table Number', 'Table Name', 'Guest Name', 'Meal Preference', 'Group']
        ws.append(headers)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='C9E8D0', end_color='C9E8D0', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for table in tables:
            for guest in sorted(table.guests, key=lambda g: g.full_name):
                ws.append([
                    table.table_number,
                    table.display_name(),
                    guest.full_name,
                    guest.meal_preference or 'Standard',
                    guest.group_name or '',
                ])
        for guest in unassigned:
            ws.append([
                '',
                'Unassigned',
                guest.full_name,
                guest.meal_preference or 'Standard',
                guest.group_name or '',
            ])
        col_widths = [14, 20, 22, 18, 20]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'seating-{wedding.partner1_name}-{wedding.partner2_name}.xlsx'.replace(' ', '-')
        response = make_response(buf.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # Default: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Table Number', 'Table Name', 'Guest Name', 'Group', 'Meal Preference'])
    for table in tables:
        for guest in sorted(table.guests, key=lambda g: g.full_name):
            writer.writerow([
                table.table_number,
                table.display_name(),
                guest.full_name,
                guest.group_name or '',
                guest.meal_preference or 'Standard',
            ])
    for guest in unassigned:
        writer.writerow([
            '',
            'Unassigned',
            guest.full_name,
            guest.group_name or '',
            guest.meal_preference or 'Standard',
        ])
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    filename = f'seating-{wedding.partner1_name}-{wedding.partner2_name}.csv'.replace(' ', '-')
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@seating_bp.route('/wedding/<int:wedding_id>/seating/bulk-assign', methods=['POST'])
@login_required
def bulk_assign(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    data = request.get_json(force=True, silent=True) or {}
    group_name = data.get('group_name')
    table_id = data.get('table_id')
    if not group_name or not table_id:
        return jsonify({'ok': False, 'error': 'Missing group_name or table_id'}), 400
    table = get_table_or_403(table_id)
    if table.wedding_id != wedding_id:
        abort(403)
    NO_GROUP_LABEL = '(No group)'
    q = Guest.query.filter_by(wedding_id=wedding_id, rsvp_status='confirmed').filter(Guest.table_id.is_(None))
    if group_name == NO_GROUP_LABEL:
        q = q.filter(db.or_(Guest.group_name.is_(None), Guest.group_name == ''))
    else:
        q = q.filter(Guest.group_name == group_name)
    unassigned = q.order_by(Guest.full_name).all()
    available = table.capacity - len(table.guests)
    assigned_ids = []
    failed = 0
    for guest in unassigned:
        if available <= 0:
            failed += 1
            continue
        guest.table_id = table.id
        guest.table_number = table.table_number
        assigned_ids.append(guest.id)
        available -= 1
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Database error'}), 500
    final_count = len(table.guests)
    return jsonify({
        'ok': True,
        'assigned': len(assigned_ids),
        'failed': failed,
        'assigned_ids': assigned_ids,
        'table_count': final_count,
    })


@seating_bp.route('/wedding/<int:wedding_id>/seating/print')
@login_required
def print_seating(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    tables = (WeddingTable.query
              .filter_by(wedding_id=wedding_id)
              .order_by(WeddingTable.table_number)
              .all())
    return render_template('wedding/seating_print.html', wedding=wedding, tables=tables)
