import csv
import io

import openpyxl
from openpyxl.styles import Font
from flask import Blueprint, flash, jsonify, redirect, render_template, url_for, request, Response
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app import db
from app.models import Guest, Wedding
from app.routes.utils import get_wedding_or_403, get_guest_or_403
from app.services.csv_service import parse_guest_csv, parse_guest_excel

guests_bp = Blueprint('guests', __name__)

_GUESTS_TAB = '#guests'

_RSVP_DB     = {'accepted': 'confirmed', 'pending': 'pending', 'declined': 'declined'}
_RSVP_LABELS = {'confirmed': 'Accepted', 'pending': 'Pending', 'declined': 'Declined'}
_RSVP_COLORS = {'confirmed': 'success',  'pending': 'warning', 'declined': 'danger'}

VALID_GROUPS = {'Bride\'s Family', 'Groom\'s Family', 'Friends', 'Colleagues', 'Other'}
VALID_MEALS  = {'Standard', 'Vegetarian', 'Vegan', 'Halal', 'Kosher', 'Other'}
VALID_RSVP   = {'pending', 'confirmed', 'declined'}

_PER_PAGE = 50


@guests_bp.route('/wedding/<int:wedding_id>/guests/search')
@login_required
def search_guests(wedding_id):
    get_wedding_or_403(wedding_id)

    q_str        = request.args.get('q',     '').strip()
    group_filter = request.args.get('group', '').strip()
    rsvp_filter  = request.args.get('rsvp',  '').strip()
    meal_filter  = request.args.get('meal',  '').strip()
    page         = request.args.get('page', 1, type=int)

    q = Guest.query.filter_by(wedding_id=wedding_id)
    if q_str:
        q = q.filter(db.or_(
            Guest.full_name.ilike(f'%{q_str}%'),
            Guest.email.ilike(f'%{q_str}%'),
        ))
    if group_filter:
        q = q.filter(Guest.group_name == group_filter)
    db_rsvp = _RSVP_DB.get(rsvp_filter)
    if db_rsvp:
        q = q.filter(Guest.rsvp_status == db_rsvp)
    if meal_filter:
        q = q.filter(Guest.meal_preference == meal_filter)

    guests_page = q.order_by(Guest.full_name).paginate(
        page=page, per_page=_PER_PAGE, error_out=False
    )

    rows_html       = render_template('wedding/_guest_rows.html',       guests_page=guests_page)
    pagination_html = render_template('wedding/_guest_pagination.html', guests_page=guests_page)

    return jsonify({
        'html':            rows_html,
        'pagination_html': pagination_html,
        'filtered_total':  guests_page.total,
        'filtered':        bool(q_str or group_filter or rsvp_filter or meal_filter),
    })


@guests_bp.route('/wedding/<int:wedding_id>/guests/add', methods=['POST'])
@login_required
def add_guest(wedding_id):
    get_wedding_or_403(wedding_id)
    detail_url = url_for('wedding.wedding_detail', wedding_id=wedding_id) + _GUESTS_TAB

    full_name       = request.form.get('full_name',       '').strip()
    email           = request.form.get('email',           '').strip() or None
    phone           = request.form.get('phone',           '').strip() or None
    group_name      = request.form.get('group_name',      '').strip() or None
    meal_preference = request.form.get('meal_preference', '').strip() or None
    rsvp_status     = request.form.get('rsvp_status',     'pending').strip()

    if not full_name:
        flash('Guest name is required.', 'danger')
        return redirect(detail_url)

    if group_name and group_name not in VALID_GROUPS:
        flash('Invalid group selection.', 'danger')
        return redirect(detail_url)

    if meal_preference and meal_preference not in VALID_MEALS:
        flash('Invalid meal preference.', 'danger')
        return redirect(detail_url)

    if rsvp_status not in VALID_RSVP:
        rsvp_status = 'pending'

    guest = Guest(
        wedding_id=wedding_id,
        full_name=full_name,
        email=email,
        phone=phone,
        group_name=group_name,
        meal_preference=meal_preference,
        rsvp_status=rsvp_status,
    )
    db.session.add(guest)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Something went wrong adding the guest. Please try again.', 'danger')
        return redirect(detail_url)

    flash(f'{full_name} has been added to the guest list.', 'success')
    return redirect(detail_url)



@guests_bp.route('/guest/<int:guest_id>/edit', methods=['POST'])
@login_required
def edit_guest(guest_id):
    guest = get_guest_or_403(guest_id)
    detail_url = url_for('wedding.wedding_detail', wedding_id=guest.wedding_id) + _GUESTS_TAB

    full_name       = request.form.get('full_name',       '').strip()
    email           = request.form.get('email',           '').strip() or None
    phone           = request.form.get('phone',           '').strip() or None
    group_name      = request.form.get('group_name',      '').strip() or None
    meal_preference = request.form.get('meal_preference', '').strip() or None
    rsvp_status     = request.form.get('rsvp_status',     'pending').strip()

    if not full_name:
        flash('Guest name is required.', 'danger')
        return redirect(detail_url)

    if group_name and group_name not in VALID_GROUPS:
        flash('Invalid group selection.', 'danger')
        return redirect(detail_url)

    if meal_preference and meal_preference not in VALID_MEALS:
        flash('Invalid meal preference.', 'danger')
        return redirect(detail_url)

    if rsvp_status not in VALID_RSVP:
        rsvp_status = 'pending'

    guest.full_name       = full_name
    guest.email           = email
    guest.phone           = phone
    guest.group_name      = group_name
    guest.meal_preference = meal_preference
    guest.rsvp_status     = rsvp_status

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Something went wrong updating the guest. Please try again.', 'danger')
        return redirect(detail_url)

    flash(f'{full_name} has been updated.', 'success')
    return redirect(detail_url)


@guests_bp.route('/wedding/<int:wedding_id>/guests/export-csv')
@login_required
def export_guests(wedding_id):
    wedding = get_wedding_or_403(wedding_id)

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=['full_name', 'email', 'phone', 'group_name', 'meal_preference', 'rsvp_status'],
        extrasaction='ignore',
        lineterminator='\r\n',
    )
    writer.writeheader()
    for guest in wedding.guests:
        writer.writerow({
            'full_name':       guest.full_name,
            'email':           guest.email        or '',
            'phone':           guest.phone        or '',
            'group_name':      guest.group_name   or '',
            'meal_preference': guest.meal_preference or '',
            'rsvp_status':     guest.rsvp_status,
        })

    filename = f"guests_{wedding.partner1_name}_{wedding.partner2_name}.csv".replace(' ', '_')
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@guests_bp.route('/wedding/<int:wedding_id>/guests/export-excel')
@login_required
def export_guests_excel(wedding_id):
    wedding = get_wedding_or_403(wedding_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Guests'

    headers = ['Full Name', 'Email', 'Phone', 'Group Name', 'Meal Preference', 'RSVP Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for guest in wedding.guests:
        ws.append([
            guest.full_name,
            guest.email           or '',
            guest.phone           or '',
            guest.group_name      or '',
            guest.meal_preference or '',
            guest.rsvp_status,
        ])

    col_widths = [20, 25, 15, 20, 18, 15]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"guests_{wedding.partner1_name}_{wedding.partner2_name}.xlsx".replace(' ', '_')
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@guests_bp.route('/wedding/<int:wedding_id>/guests/download-excel-template')
@login_required
def download_excel_template(wedding_id):
    get_wedding_or_403(wedding_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Guests'

    headers = ['Full Name', 'Email', 'Phone', 'Group Name', 'Meal Preference', 'RSVP Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.append(['John Smith', 'john@example.com', '+1555010101', "Groom's Family", 'Standard', 'pending'])
    ws.append(['Jane Doe',   'jane@example.com', '+1555010102', "Bride's Family",  'Vegetarian', 'accepted'])

    col_widths = [20, 25, 15, 20, 18, 15]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="guest_list_template.xlsx"'},
    )


@guests_bp.route('/wedding/<int:wedding_id>/guests/import-csv', methods=['POST'])
@login_required
def import_guests(wedding_id):
    get_wedding_or_403(wedding_id)
    detail_url = url_for('wedding.wedding_detail', wedding_id=wedding_id) + _GUESTS_TAB

    guest_file = request.files.get('guest_file')
    if not guest_file or guest_file.filename == '':
        flash('Please select a file to upload.', 'danger')
        return redirect(detail_url)

    filename = secure_filename(guest_file.filename).lower()
    if filename.endswith('.csv'):
        guests, errors = parse_guest_csv(guest_file)
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        guests, errors = parse_guest_excel(guest_file)
    else:
        flash('Please upload a CSV or Excel file (.csv, .xlsx).', 'danger')
        return redirect(detail_url)

    if not guests and errors:
        for msg in errors:
            flash(msg, 'danger')
        return redirect(detail_url)

    # Build a set of existing guest names for this wedding (case-insensitive)
    existing_names = {
        g.full_name.lower()
        for g in Guest.query.filter_by(wedding_id=wedding_id).with_entities(Guest.full_name)
    }

    added = 0
    seen_in_csv = set()
    for data in guests:
        name_key = data['full_name'].lower()
        if name_key in existing_names or name_key in seen_in_csv:
            errors.append(f'Skipped duplicate: {data["full_name"]}')
            continue
        seen_in_csv.add(name_key)
        db.session.add(Guest(wedding_id=wedding_id, **data))
        added += 1

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Something went wrong saving the imported guests. Please try again.', 'danger')
        return redirect(detail_url)

    imported = added
    skipped  = len(errors)
    summary  = f'{imported} guest{"s" if imported != 1 else ""} imported'
    if skipped:
        summary += f', {skipped} row{"s" if skipped > 1 else ""} skipped'
    summary += '.'
    flash(summary, 'success' if not skipped else 'warning')

    for msg in errors:
        flash(msg, 'warning')

    return redirect(detail_url)


@guests_bp.route('/guest/<int:guest_id>/update-rsvp', methods=['POST'])
@login_required
def update_rsvp(guest_id):
    guest = get_guest_or_403(guest_id)
    data = request.get_json(silent=True) or {}
    db_status = _RSVP_DB.get(data.get('rsvp_status', '').lower())
    if not db_status:
        return {'error': 'Invalid RSVP status'}, 400
    guest.rsvp_status = db_status
    try:
        db.session.commit()
        return {'ok': True, 'db_status': db_status,
                'label': _RSVP_LABELS[db_status], 'color': _RSVP_COLORS[db_status]}
    except Exception:
        db.session.rollback()
        return {'error': 'Could not update'}, 500


@guests_bp.route('/wedding/<int:wedding_id>/guests/bulk-update', methods=['POST'])
@login_required
def bulk_update_guests(wedding_id):
    get_wedding_or_403(wedding_id)
    data      = request.get_json(silent=True) or {}
    guest_ids = [int(i) for i in data.get('guest_ids', []) if str(i).isdigit()]
    action    = data.get('action', '')

    if not guest_ids:
        return {'error': 'No guests selected'}, 400

    base = Guest.query.filter(Guest.id.in_(guest_ids), Guest.wedding_id == wedding_id)

    if action.startswith('rsvp_'):
        db_status = _RSVP_DB.get(action[5:])
        if not db_status:
            return {'error': 'Invalid action'}, 400
        base.update({'rsvp_status': db_status}, synchronize_session=False)
    elif action == 'delete':
        base.delete(synchronize_session=False)
    else:
        return {'error': 'Invalid action'}, 400

    try:
        db.session.commit()
        return {'ok': True}
    except Exception:
        db.session.rollback()
        return {'error': 'Could not complete action'}, 500


@guests_bp.route('/wedding/<int:wedding_id>/guests/delete-selected', methods=['POST'])
@login_required
def delete_selected_guests(wedding_id):
    get_wedding_or_403(wedding_id)
    detail_url = url_for('wedding.wedding_detail', wedding_id=wedding_id) + _GUESTS_TAB

    raw_ids = request.form.getlist('guest_ids')
    ids = [int(i) for i in raw_ids if i.isdigit()]

    if not ids:
        flash('No guests selected.', 'warning')
        return redirect(detail_url)

    count = Guest.query.filter(
        Guest.id.in_(ids),
        Guest.wedding_id == wedding_id,
    ).delete(synchronize_session=False)

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Something went wrong. Please try again.', 'danger')
        return redirect(detail_url)

    flash(f'{count} guest{"s" if count != 1 else ""} deleted.', 'success')
    return redirect(detail_url)


@guests_bp.route('/wedding/<int:wedding_id>/guests/deduplicate', methods=['POST'])
@login_required
def deduplicate_guests(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    detail_url = url_for('wedding.wedding_detail', wedding_id=wedding_id) + _GUESTS_TAB

    seen = {}
    to_delete = []
    for guest in wedding.guests:
        key = guest.full_name.lower()
        if key in seen:
            to_delete.append(guest)
        else:
            seen[key] = guest

    if not to_delete:
        flash('No duplicate guests found.', 'info')
        return redirect(detail_url)

    try:
        for guest in to_delete:
            db.session.delete(guest)
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Something went wrong removing duplicates. Please try again.', 'danger')
        return redirect(detail_url)

    count = len(to_delete)
    flash(f'{count} duplicate guest{"s" if count != 1 else ""} removed.', 'success')
    return redirect(detail_url)


@guests_bp.route('/wedding/<int:wedding_id>/guests/delete-all', methods=['POST'])
@login_required
def delete_all_guests(wedding_id):
    wedding = get_wedding_or_403(wedding_id)
    detail_url = url_for('wedding.wedding_detail', wedding_id=wedding_id) + _GUESTS_TAB

    count = len(wedding.guests)
    if count == 0:
        flash('There are no guests to delete.', 'warning')
        return redirect(detail_url)

    try:
        Guest.query.filter_by(wedding_id=wedding_id).delete()
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Something went wrong deleting all guests. Please try again.', 'danger')
        return redirect(detail_url)

    flash(f'All {count} guest{"s" if count != 1 else ""} removed.', 'success')
    return redirect(detail_url)


@guests_bp.route('/guest/<int:guest_id>/delete', methods=['POST'])
@login_required
def delete_guest(guest_id):
    guest = get_guest_or_403(guest_id)
    detail_url = url_for('wedding.wedding_detail', wedding_id=guest.wedding_id) + _GUESTS_TAB

    name = guest.full_name
    try:
        db.session.delete(guest)
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Something went wrong deleting the guest. Please try again.', 'danger')
        return redirect(detail_url)

    flash(f'{name} has been removed from the guest list.', 'success')
    return redirect(detail_url)
